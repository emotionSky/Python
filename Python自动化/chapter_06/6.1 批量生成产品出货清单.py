from openpyxl import load_workbook  # pip install openpyxl --user
from pathlib import Path

dst_folder = Path('./输出表格')
if not dst_folder.exists():
    dst_folder.mkdir(parents=True)

workbook = load_workbook('./素材/出货统计表.xlsx')
worksheet = workbook['Sheet1']
data = {}

# 行从1开始，第一行是表头，所以需要从2开始
for row in range(2, worksheet.max_row + 1):
    date = worksheet['B' + str(row)].value.date()
    customer = worksheet['C' + str(row)].value
    product = worksheet['D' + str(row)].value
    number = worksheet['E' + str(row)].value
    model = worksheet['G' + str(row)].value
    info_list = [customer, product, number, model]
    data.setdefault(date, [])
    data[date].append(info_list)

for key, value in data.items():
    print(key, value)

workbook_day = load_workbook('./素材/出货清单模板.xlsx')
worksheet_day = workbook_day['出货清单模板']
for date in data.keys():
    worksheet_new = workbook_day.copy_worksheet(worksheet_day)  # 拷贝工作表来新建
    worksheet_new.title = str(date)[-5:]                        # 使用出货日期中的月和日来命名Sheet
    worksheet_new.cell(row=2, column=5).value = date            # 将出货日期写在 2行5列
    i = 4  # 出货内容从第4行开始
    for product in data[date]:
        worksheet_new.cell(row=i, column=2).value = product[0]
        worksheet_new.cell(row=i, column=3).value = product[1]
        worksheet_new.cell(row=i, column=4).value = product[2]
        worksheet_new.cell(row=i, column=5).value = product[3]
        i += 1

dst_file = dst_folder / '产品出货清单.xlsx'
workbook_day.save(str(dst_file))
